# Importação das bibliotecas necessárias
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

# Configuração do WebDriver para usar o Chrome
driver = webdriver.Chrome()
# Acessa a página de ofertas do dia do site Kabum
driver.get('https://www.kabum.com.br/ofertas/ofertadodia?pagina=1')

# Extrai todos os títulos dos produtos usando XPATH
titulos = driver.find_elements(By.XPATH, "//span[@class='sc-d79c9c3f-0 nlmfp sc-cdc9b13f-16 eHyEuD nameCard']")

# Extrai todos os preços originais (De R$) dos produtos usando o XPATH fornecido
precos_de = driver.find_elements(By.XPATH, "//span[@class='sc-620f2d27-1 bksuMM oldPriceCard']")

# Extrai todos os preços promocionais (Por R$) dos produtos usando o XPATH fornecido
precos_por = driver.find_elements(By.XPATH, "//span[@class='sc-620f2d27-2 bMHwXA priceCard']")

# Extrai os URLs dos produtos
urls = driver.find_elements(By.XPATH, "//a[contains(@class,'productLink')]")

# Cria uma nova planilha Excel e remove a aba padrão
workbook = openpyxl.Workbook()
sheet_default = workbook.active
workbook.remove(sheet_default)
# Adiciona uma nova aba chamada 'Produtos' e a seleciona para manipulação
sheet_produtos = workbook.create_sheet('Produtos')

# Configura os títulos das colunas na primeira linha da planilha
sheet_produtos['A1'].value = 'Produto'
sheet_produtos['B1'].value = 'De R$'
sheet_produtos['C1'].value = 'Por R$'
sheet_produtos['D1'].value = 'Desconto'

# Itera sobre os títulos, preços originais, preços promocionais dos produtos e URLs extraídos
for titulo, preco_de, preco_por, url in zip(titulos, precos_de, precos_por, urls):
    # Remove caracteres não numéricos para conversão em float dos preços
    preco_de_num = float(preco_de.text.replace('R$', '').replace('.', '').replace(',', '.').strip())
    preco_por_num = float(preco_por.text.replace('R$', '').replace('.', '').replace(',', '.').strip())
    
    # Calcula a porcentagem de desconto
    desconto = ((preco_de_num - preco_por_num) / preco_de_num) * 100
    
    # Adiciona os dados na próxima linha disponível da planilha
    row = [titulo.text, preco_de.text, preco_por.text, f'{desconto:.2f}%']
    sheet_produtos.append(row)
    # Define o hyperlink para o título do produto
    cell = sheet_produtos.cell(row=sheet_produtos.max_row, column=1)
    cell.hyperlink = url.get_attribute('href')
    cell.style = 'Hyperlink'

# Salva a planilha no disco com o nome 'Ofertas do dia.xlsx'
workbook.save('Ofertas do dia.xlsx')

# Fecha o navegador
driver.quit()

